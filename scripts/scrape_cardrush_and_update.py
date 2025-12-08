# -*- coding: utf-8 -*-
"""
CardRush(ポケカ)をスクレイピングして最新買取リストを取得し、
ポケカラッシュ.xlsx の Sheet1（Myca商品マスタ形式）と照合して値付け、
・ポケカラッシュ_一致抽出.xlsx に一致データ＆レポートを出力
・ポケカラッシュ.xlsx の Sheet1 に 1〜5行を残したまま 6行目から上書き
・Mycaアップロード用CSVを自動出力
まで一括で行うスクリプト。
"""

import os
import sys
import re
import time
import unicodedata
import math
import tempfile
import shutil

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pathlib import Path 

# ================== 設定 ==================
# ★ GitHubリポジトリ(climax3) 基準のパスに変更
REPO_ROOT = Path(__file__).resolve().parent.parent
BASE_DIR  = REPO_ROOT / "data"

# .xlsm に変更（さっき data に入れたファイル名に合わせる）
XLSX_FILE       = BASE_DIR / "pokeca_rush.xlsm"             # 元のMyca形式のファイル
MATCH_OUT_FILE  = BASE_DIR / "ポケカラッシュ_一致抽出.xlsx"      # 一致抽出＆レポート出力用
MYCA_CSV_FILE   = BASE_DIR / "ポケカラッシュ_Mycaアップロード用.csv"


SHEET1_NAME = "Sheet1"

# Sheet1 側の列位置（Mycaテンプレそのまま）
S1_NAME_COL_LETTER  = "C"  # 名前 (display_name)
S1_MODEL_COL_LETTER = "F"  # 型番 (cardnumber)
S1_PRICE_COL_LETTER = "O"  # 価格（ここを上書き = buy_price）
S1_G_COL_LETTER     = "G"  # rarity (AR 判定用)

# CardRush 側（スクレイピング結果 DataFrame）を
# A=名前 / B=型番 / C=価格 として扱う
S2_NAME_COL_LETTER  = "A"
S2_MODEL_COL_LETTER = "B"
S2_PRICE_COL_LETTER = "C"

threshold = 0.90  # 名前＋型番一致率の平均がこれ以上なら同一カードとみなす

# ====== スクレイピング設定 ======
BASE_URL = "https://cardrush.media/pokemon/buying_prices"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    )
}
PRICE_THRESHOLD = 90  # この金額以下は除外

# 型番抽出パターン
MODEL_CANDIDATE_RE = re.compile(r"[A-Za-z0-9]+\/[A-Za-z0-9]+")
ALT_MODEL_RE = re.compile(r"[A-Za-z0-9]{1,6}[-\/][A-Za-z0-9]{1,10}(?:\/[0-9]{1,6})?")


# ================== 共通ユーティリティ ==================
try:
    from rapidfuzz.distance import Levenshtein

    def sim_ratio(a, b):
        if not a and not b:
            return 1.0
        if not a or not b:
            return 0.0
        return Levenshtein.normalized_similarity(a, b)

except Exception:
    from difflib import SequenceMatcher

    def sim_ratio(a, b):
        if not a and not b:
            return 1.0
        if not a or not b:
            return 0.0
        return SequenceMatcher(None, a, b).ratio()


def get_col_by_letter(df, letter: str) -> str:
    """A,B,C... から DataFrame の列名を取得"""
    idx = ord(letter.upper()) - ord("A")
    if idx < 0 or idx >= len(df.columns):
        raise IndexError(f"列 {letter} がシートに存在しません（列数={len(df.columns)}）")
    return df.columns[idx]


def normalize_key(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s)).lower()
    for ch in [
        "（", "）", "[", "]", "(", ")", "「", "」", "『", "』", "【", "】",
        "★", "☆", "・", "/", "\\", "-", "_", "　", " ", "：", ":"
    ]:
        s = s.replace(ch, "")
    return s


def frag(s, n=3):
    return s[:n] if s else ""


def try_parse_price(text: str):
    if not text:
        return None
    s = text.replace("¥", "").replace("円", "").replace(",", "").strip()
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


# ================== スクレイピング部 ==================
def detect_columns(table):
    """テーブルのヘッダがあればインデックスを推定。無ければデフォルト"""
    headers = []
    thead = table.find("thead")
    if thead:
        headers = [th.get_text(strip=True) for th in thead.find_all("th")]
    else:
        first_tr = table.find("tr")
        if first_tr:
            ths = first_tr.find_all("th")
            if ths:
                headers = [th.get_text(strip=True) for th in ths]

    name_idx = model_idx = price_idx = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(k in hl for k in ("カード", "商品", "name")):
            name_idx = i
        if any(k in hl for k in ("型番", "型", "code", "model", "品番", "番号")):
            model_idx = i
        if any(k in hl for k in ("買取", "買", "価格", "円", "price")):
            price_idx = i

    if name_idx is None:
        name_idx = 0
    if model_idx is None:
        model_idx = 3
    if price_idx is None:
        price_idx = -1
    return name_idx, model_idx, price_idx


def parse_page(html: str):
    """1ページ分のHTMLを解析し、カード情報リストを返す"""
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if not table:
        return []

    name_idx, model_idx, price_idx = detect_columns(table)
    rows = table.find_all("tr")
    results = []

    for tr in rows:
        tds = tr.find_all(["td", "th"])
        if not tds:
            continue
        cells = [td.get_text(" ", strip=True) for td in tds]

        # ヘッダ行っぽいものはスキップ
        joined = " ".join(cells)
        if ("カード" in joined and "買取" in joined) or ("商品" in joined and "価格" in joined):
            continue

        name = cells[name_idx] if name_idx < len(cells) else ""
        model = cells[model_idx] if model_idx < len(cells) else ""

        # 型番が空ならセル全体から抽出
        if not model:
            for c in cells:
                m = MODEL_CANDIDATE_RE.search(c)
                if m:
                    model = m.group()
                    break
            if not model:
                m2 = ALT_MODEL_RE.search(" ".join(cells))
                if m2:
                    model = m2.group()

        # 価格抽出
        if 0 <= price_idx < len(cells):
            price = try_parse_price(cells[price_idx])
        else:
            price = try_parse_price(cells[-1] if cells else "")

        if price is None or price <= PRICE_THRESHOLD:
            continue

        results.append(
            {
                "name": name.strip(),
                "model": model.strip(),
                "price": price,
            }
        )
    return results


def fetch_page(page: int, session: requests.Session):
    """ページを取得"""
    url = BASE_URL if page == 1 else f"{BASE_URL}?page={page}"
    resp = session.get(url, headers=HEADERS, timeout=20)
    if resp.status_code != 200:
        print(f"⚠️ HTTP {resp.status_code}: {url}")
        return None
    return resp.text


def scrape_cardrush():
    """CardRush 全ページをスクレイピングして DataFrame を返す"""
    all_cards = []
    page = 1

    with requests.Session() as session:
        while True:
            print(f"📄 Fetching page {page}...")
            html = fetch_page(page, session)
            if not html:
                print("❌ HTTPエラーまたは空ページ。停止します。")
                break

            items = parse_page(html)
            if not items:
                print("🔚 このページに抽出項目がありません。終了します。")
                break

            all_cards.extend(items)
            print(f"✅ Page {page}: {len(items)}件取得。合計 {len(all_cards)}件。")

            page += 1
            time.sleep(0.3)  # サーバー負荷軽減

    if not all_cards:
        raise RuntimeError("⚠️ 抽出結果が空です。")

    df = pd.DataFrame(all_cards).drop_duplicates(subset=["name", "model", "price"])
    df = df[["name", "model", "price"]]  # A:名前, B:型番, C:価格
    return df


# ================== 値付けルール ==================
def adjust_price(row, s1_price_col, g_col_name):
    """
    CardRush側の価格(S2_照合価格) と rarity(G列)から、
    最終的に Sheet1 の O列へ入れる価格を計算する。
    """

    raw = row.get("S2_照合価格", None)
    fallback = row.get(s1_price_col, None)

    # S2側価格がない → 元の価格のまま
    if pd.isna(raw) or raw is None or str(raw).strip() == "":
        return fallback

    s = str(raw).replace(",", "").strip()
    try:
        base = float(s)
    except ValueError:
        return fallback

    p = int(base)
    g_val = str(row.get(g_col_name, "") or "")
    new_p = p

    # --- 個別マッピング ---
    special_map = {
        100: 50, 150: 50,
        200: 100,
        300: 150,
        400: 200,
        500: 300,
        600: 400,
        700: 500,
        800: 600,
        900: 700,
        1000: 800,
    }
    if p in special_map:
        new_p = special_map[p]

    # --- 10000以上の基本ルール ---
    elif p >= 10000:
        # まず1.05倍
        tmp = int(p * 1.05)

        if tmp >= 40000:
            # 4万以上は無条件で1000単位切り上げ
            new_p = ((tmp // 1000) + 1) * 1000
        else:
            # 10000〜39999 → 100の位で分岐
            hundred = (tmp // 100) % 10

            # 1〜5 → 1000円単位で切り捨て
            if 1 <= hundred <= 5:
                new_p = (tmp // 1000) * 1000
            # 6〜9 → 1000円単位で切り上げ
            elif 6 <= hundred <= 9:
                new_p = ((tmp // 1000) + 1) * 1000
            # 0 → そのまま1000円単位
            else:
                new_p = (tmp // 1000) * 1000

    # 1001〜9999 → 変更なし
    # 0〜99 も特にルールなし

    # --- ここから最終仕上げルール ---

    # 6桁（100,000〜999,999）は 1万の位で四捨五入
    if 100000 <= new_p <= 999999:
        q = new_p // 10000      # 万の位
        r = new_p % 10000       # 下4桁
        if r >= 5000:
            q += 1
        new_p = q * 10000

    # 7桁以上（1,000,000〜）は 10万の位を常に切り上げ
    if new_p >= 1000000:
        new_p = ((new_p + 100000 - 1) // 100000) * 100000

    # --- AR補正 ---
    if new_p <= 299 and g_val == "AR":
        new_p = 300

    return new_p


# ================== メイン処理 ==================
def main():
    xlsx_path     = XLSX_FILE
    match_out_path = MATCH_OUT_FILE
    myca_csv_path  = MYCA_CSV_FILE

    if not os.path.exists(xlsx_path):
        print(f"✖ 入力ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    # ---- ① CardRush スクレイピング（Sheet2相当） ----
    print("=== CardRush スクレイピング開始 ===")
    s2 = scrape_cardrush()
    print(f"=== スクレイピング完了: {len(s2)} 件 ===")

    # DataFrame の列は [name, model, price] なので
    s2.columns = ["name", "model", "price"]

    # ---- ② 元ファイルの Sheet1 を読み込み ----
    s1 = pd.read_excel(xlsx_path, sheet_name=SHEET1_NAME, dtype=str)

    # 列取得
    s1_name_col  = get_col_by_letter(s1, S1_NAME_COL_LETTER)
    s1_model_col = get_col_by_letter(s1, S1_MODEL_COL_LETTER)
    s1_price_col = get_col_by_letter(s1, S1_PRICE_COL_LETTER)
    s1_g_col     = get_col_by_letter(s1, S1_G_COL_LETTER)  # AR判定

    s2_name_col  = get_col_by_letter(s2, S2_NAME_COL_LETTER)  # "name"
    s2_model_col = get_col_by_letter(s2, S2_MODEL_COL_LETTER) # "model"
    s2_price_col = get_col_by_letter(s2, S2_PRICE_COL_LETTER) # "price"

    # 正規化キー
    s1["_name"]  = s1[s1_name_col].fillna("").map(normalize_key)
    s1["_model"] = s1[s1_model_col].fillna("").map(normalize_key)
    s2["_name"]  = s2[s2_name_col].fillna("").map(normalize_key)
    s2["_model"] = s2[s2_model_col].fillna("").map(normalize_key)

    # バケット作成
    s2["__k_model3"] = s2["_model"].map(lambda x: frag(x, 3))
    s2["__k_name3"]  = s2["_name"].map(lambda x: frag(x, 3))

    bucket_full, bucket_model, bucket_name = {}, {}, {}
    for idx, row in s2.iterrows():
        bucket_full.setdefault((row["__k_model3"], row["__k_name3"]), []).append(idx)
        bucket_model.setdefault(row["__k_model3"], []).append(idx)
        bucket_name.setdefault(row["__k_name3"], []).append(idx)

    # ---- ③ 照合（同じ s2 行をなるべく再利用しない）----
    rows = []
    used_s2 = set()

    for i, r1 in s1.iterrows():
        k = (frag(r1["_model"], 3), frag(r1["_name"], 3))

        # 基本候補
        base_candidates = (
            bucket_full.get(k)
            or bucket_model.get(k[0])
            or bucket_name.get(k[1])
            or list(s2.index)
        )

        # ★「ひかる～」専用絞り込みロジック
        orig_name_s1 = str(r1[s1_name_col] or "")

        if "ひかる" in orig_name_s1:
            cand = [
                j for j in base_candidates
                if "ひかる" in str(s2.at[j, s2_name_col] or "")
            ]

            if "旧" in orig_name_s1:
                cand_kyu = [
                    j for j in cand
                    if "旧" in str(s2.at[j, s2_name_col] or "")
                ]
                if cand_kyu:
                    cand = cand_kyu

            candidates = cand if cand else base_candidates
        else:
            candidates = base_candidates

        best_j, best_score, best_nr, best_mr = None, -1, 0, 0
        best_unused_j, best_unused_score = None, -1
        best_unused_nr, best_unused_mr = 0, 0

        n1, m1 = r1["_name"], r1["_model"]

        for j in candidates:
            n2, m2 = s2.at[j, "_name"], s2.at[j, "_model"]
            nr, mr = sim_ratio(n1, n2), sim_ratio(m1, m2)
            sc = (nr + mr) / 2

            # 全体のベスト
            if sc > best_score:
                best_score = sc
                best_j = j
                best_nr, best_mr = nr, mr

            # 未使用の中でのベスト
            if j not in used_s2 and sc > best_unused_score:
                best_unused_score = sc
                best_unused_j = j
                best_unused_nr, best_unused_mr = nr, mr

            if sc >= 0.999:
                # ほぼ完全一致ならそれ以上見ない
                continue

        # 未使用で閾値以上の候補があればそっち優先
        if best_unused_j is not None and best_unused_score >= threshold:
            chosen_j = best_unused_j
            chosen_score = best_unused_score
            chosen_nr, chosen_mr = best_unused_nr, best_unused_mr
        else:
            chosen_j = best_j
            chosen_score = best_score
            chosen_nr, chosen_mr = best_nr, best_mr

        if chosen_j is not None:
            used_s2.add(chosen_j)

        rows.append({
            "s1_idx": i,
            "s2_idx": chosen_j,
            "名前一致率": chosen_nr,
            "型番一致率": chosen_mr,
            "平均一致率": chosen_score,
            "閾値以上": chosen_score >= threshold
        })

    match_df = pd.DataFrame(rows)

    # ---- ④ レポート作成 ----
    s1_key = s1[[s1_name_col, s1_model_col, s1_price_col]].copy()
    s1_key.columns = ["S1_名前", "S1_型番", "S1_価格"]
    s1_key["s1_idx"] = s1.index

    s2_key = s2[[s2_name_col, s2_model_col, s2_price_col]].copy()
    s2_key.columns = ["S2_名前", "S2_型番", "S2_価格"]
    s2_key["s2_idx"] = s2.index

    report = (
        match_df
        .merge(s1_key, on="s1_idx", how="left")
        .merge(s2_key, on="s2_idx", how="left")
        [["S1_名前","S1_型番","S1_価格",
          "S2_名前","S2_型番","S2_価格",
          "名前一致率","型番一致率","平均一致率","閾値以上"]]
    )

    # ---- ⑤ 一致した行だけ抽出 ----
    keep_idx = match_df.loc[match_df["閾値以上"], "s1_idx"]
    if len(keep_idx) == 0:
        print("⚠ 一致した行が1件もありません（threshold を下げるとマッチするかも）")
        sys.exit(0)

    s1_filtered = s1.loc[keep_idx].copy().reset_index(drop=True)
    s1_filtered["s1_idx"] = keep_idx.values

    # s2側情報を結合
    s2_info = (
        match_df.loc[match_df["閾値以上"], ["s1_idx","s2_idx","平均一致率"]]
        .merge(s2_key, on="s2_idx", how="left")
        .rename(columns={
            "S2_価格": "S2_照合価格",
            "S2_名前": "S2_照合名前",
            "S2_型番": "S2_照合型番"
        })
    )

    s1_filtered = s1_filtered.merge(
        s2_info[["s1_idx","S2_照合価格","S2_照合名前","S2_照合型番","平均一致率"]],
        on="s1_idx",
        how="left"
    )

    # ---- ⑥ 価格変換 ----
    s1_filtered["__new_price"] = s1_filtered.apply(
        lambda r: adjust_price(r, s1_price_col, s1_g_col),
        axis=1
    )

    # O列へ上書き（DataFrame上）
    s1_filtered[s1_price_col] = s1_filtered["__new_price"]
    s1_filtered.drop(columns=["__new_price"], inplace=True)

    # ---- ⑦ 一致抽出ファイル（確認用）を保存 ----
    match_out_full = match_out_path
    with pd.ExcelWriter(match_out_full, engine="openpyxl") as writer:
        s1_filtered.to_excel(writer, sheet_name="Sheet1_一致抽出", index=False)
        report.to_excel(writer, sheet_name="照合レポート", index=False)
    print(f"✓ 一致抽出ファイル出力: {match_out_full}")

    # === ここから、「ポケカラッシュ.xlsx の Sheet1 に 6行目から上書き」 ===

    # 元のSheet1の列構成を取得（s1_filteredには余計な列も入ってるので）
    original_columns = s1.columns  # 元Sheet1の列順
    export_df = s1_filtered[original_columns].copy()

    # ---- ⑧ Excelブックを開いて Sheet1 を更新 ----
    wb = load_workbook(xlsx_path, keep_vba=True)  # ここを変更
    ws = wb[SHEET1_NAME]


    # 1〜5行目はそのまま残し、6行目から export_df を書き込む
    start_row = 6
    n_rows, n_cols = export_df.shape

    # まず、既存の 6行目以降を全部クリアしておく（見た目を綺麗にするため）
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

    # 6行目から書き込み（ヘッダー行は不要なので、DataFrameの内容だけ）
    for i in range(n_rows):
        row_idx = start_row + i
        for j in range(n_cols):
            value = export_df.iat[i, j]
            ws.cell(row=row_idx, column=j + 1).value = value

    wb.save(xlsx_path)
    print(f"✓ ポケカラッシュ.xlsx の Sheet1 を 6行目から上書きしました: {xlsx_path}")

    # ---- ⑨ Myca用CSV自動出力 ----
    # 上書き後の Sheet1 をそのままCSV化（1〜5行目も含めて）
    df_for_csv = pd.read_excel(xlsx_path, sheet_name=SHEET1_NAME, header=None, dtype=object)
    df_for_csv.to_csv(myca_csv_path, index=False, header=False, encoding="utf-8-sig")

    print(f"✓ Mycaアップロード用CSVを出力しました: {myca_csv_path}")
    print("=== 全処理完了 ===")


if __name__ == "__main__":
    main()
